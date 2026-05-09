#!/usr/bin/env python3
"""
Sync locally-provided digital presentation images into the Cannesdy spreadsheet.

Walks digital-presentation-images/originals/, parses each filename's
category-brand-campaign convention, converts to web-optimized WebP at q90 (max
2048px on the long edge), writes the .webp into digital-presentation-images/,
and updates the matching row(s)' Image URL column in the xlsx.

Run with --dry-run first to preview matches and conversions without writing.

Filename convention:
    <category-slug>-<brand-words>-<campaign-words>.<ext>

Where <category-slug> matches one of the SHEET_BY_PREFIX entries below.
"""
from __future__ import annotations

import argparse
import os
import re
import shutil
import sys
import unicodedata
from collections import defaultdict
from pathlib import Path

from PIL import Image
import openpyxl

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "Brain_Cannesdy_Brand_Experience_2025.xlsx"
ORIGINALS = ROOT / "digital-presentation-images" / "originals"
WEB_DIR = ROOT / "digital-presentation-images"
REL_PREFIX = "digital-presentation-images/"

# Map category-slug-prefix in filenames -> exact xlsx sheet name.
# Add new categories here as they come in.
SHEET_BY_PREFIX = {
    "audio-and-radio": "Audio & Radio",
    "brand-experience-and-activation": "Brand Experience & Activation",
    "creative-business-transformation": "Creative Bus. Transformation",
    "creative-effectiveness": "Creative Effectiveness",
    "creative-strategy": "Creative Strategy",
    "direct": "Direct",
    "entertainment": "Entertainment",
    "entertainment-lions-for-gaming": "Entertainment Lions for Gaming",
    "entertainment-lions-for-music": "Entertainment Lions for Music",
    "entertainment-lions-for-sport": "Entertainment Lions for Sport",
    "film": "Film",
    "film-craft": "Film Craft",
    "grand-prix-for-good": "Grand Prix for Good",
    "health-and-wellness": "Health & Wellness",
    "industry-craft": "Industry Craft",
    "lions-health-grand-prix-for-good": "Health Grand Prix for Good",
    "media": "Media",
    "outdoor": "Outdoor",
    "pharma": "Pharma",
    "pr": "PR",
    "print-and-publishing": "Print & Publishing",
    "social-and-creator": "Social & Creator",
    "sustainable-development-goals": "Sustainable Development Goals",
}

# Match longer prefixes first so e.g. "entertainment-lions-for-music" wins over "entertainment".
PREFIXES_BY_LENGTH = sorted(SHEET_BY_PREFIX.keys(), key=len, reverse=True)

WEBP_QUALITY = 80  # q80 keeps small text crisp on WebP and saves ~20% vs originals
MAX_DIM = 2048

# ---------------------------------------------------------------------------
# Normalization
# ---------------------------------------------------------------------------

def slugify(value: str) -> str:
    """Lowercase, fold diacritics, drop apostrophes, hyphenate the rest."""
    if value is None:
        return ""
    s = str(value)
    # Fold diacritics: 'unión' -> 'union', 'gavião' -> 'gaviao', 'raça' -> 'raca'.
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = s.lower()
    # Expand ampersand into a word so it survives the alphanumeric pass.
    s = s.replace("&", " and ")
    # Drop apostrophe-like characters entirely so "Senna's" matches "sennas",
    # rather than splitting into "senna" and "s" tokens.
    s = re.sub(r"[’'`‘]", "", s)
    s = re.sub(r"[^a-z0-9]+", "-", s)
    return s.strip("-")


def tokens(slug: str) -> set[str]:
    return {t for t in slug.split("-") if t}


def jaccard(a: set[str], b: set[str]) -> float:
    if not a or not b:
        return 0.0
    return len(a & b) / len(a | b)


# ---------------------------------------------------------------------------
# Filename parsing
# ---------------------------------------------------------------------------

def parse_filename(stem: str) -> tuple[str | None, str]:
    """Return (sheet_name, match_key) where match_key is the brand+campaign portion."""
    for prefix in PREFIXES_BY_LENGTH:
        if stem == prefix or stem.startswith(prefix + "-"):
            sheet = SHEET_BY_PREFIX[prefix]
            rest = stem[len(prefix) + 1 :] if stem != prefix else ""
            return sheet, rest
    return None, stem


# ---------------------------------------------------------------------------
# Row matching
# ---------------------------------------------------------------------------

def match_rows(file_key: str, sheet_rows: list[dict]) -> list[dict]:
    """
    Find xlsx rows matching the file_key.

    Strategy: score every row by the relationship between file_key tokens and the
    row's brand+campaign tokens. Return all rows tied at the top score above the
    threshold. This naturally handles three cases:

      1. Same campaign / multiple subcat wins: all tied at score=1.0 (exact match).
      2. Per-piece variants (Bundles of Joy OOH1 vs OOH2): the one with matching
         suffix wins because its token set overlap is highest.
      3. Brand-name variations (Burger King vs Burger King UK): subset relationship
         keeps Jaccard high enough to clear threshold.
    """
    file_tokens = tokens(file_key)
    if not file_tokens:
        return []

    scored = []
    for row in sheet_rows:
        row_key = slugify(f"{row['brand']} {row['campaign']}")
        row_tokens = tokens(row_key)
        if not row_tokens:
            continue

        # Score: prioritise exact, then subset, then jaccard.
        if row_key == file_key:
            score = 1.0
        elif file_key in row_key or row_key in file_key:
            # Subset case (e.g., file "burger-king-..." and row "burger-king-uk-...").
            # Compute jaccard but boost slightly so subsets beat partial matches.
            score = 0.5 + 0.5 * jaccard(file_tokens, row_tokens)
        else:
            score = jaccard(file_tokens, row_tokens)
        scored.append((score, row))

    if not scored:
        return []
    top = max(s for s, _ in scored)
    if top < 0.60:  # threshold below which we'd rather report unmatched
        return []
    return [row for s, row in scored if s == top]


# ---------------------------------------------------------------------------
# Image conversion
# ---------------------------------------------------------------------------

def convert_to_webp(src: Path, dst: Path, force: bool = False) -> tuple[int, int, int, bool]:
    """Convert src image -> dst webp. Skip if dst exists and is newer than src.

    Returns (orig_kb, new_kb, longest_dim, was_skipped).
    """
    if not force and dst.exists() and dst.stat().st_mtime >= src.stat().st_mtime:
        # Already converted; skip the slow re-encode — but verify the dst
        # is actually readable, in case a previous run was interrupted mid-write.
        try:
            with Image.open(dst) as im:
                im.verify()
            with Image.open(dst) as im:
                return (os.path.getsize(src) // 1024,
                        os.path.getsize(dst) // 1024,
                        max(im.size), True)
        except Exception:
            # Corrupt/partial output — fall through and re-encode.
            pass
    im = Image.open(src)
    # Resize so the longest edge is <= MAX_DIM.
    w, h = im.size
    scale = min(1.0, MAX_DIM / max(w, h))
    if scale < 1.0:
        new_size = (round(w * scale), round(h * scale))
        im = im.resize(new_size, Image.LANCZOS)
    # Convert palette/CMYK to RGB to keep webp encoding sane.
    if im.mode not in ("RGB", "RGBA"):
        im = im.convert("RGB")
    dst.parent.mkdir(parents=True, exist_ok=True)
    im.save(dst, format="WEBP", quality=WEBP_QUALITY, method=6)
    return os.path.getsize(src) // 1024, os.path.getsize(dst) // 1024, max(im.size), False


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def load_sheet_rows(ws) -> list[dict]:
    headers = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}
    rows = []
    for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        values = [c.value for c in row]
        if values[0] is None:
            continue
        rows.append({
            "row_idx": r_idx,
            "id": values[idx["#"]],
            "tier": values[idx["Tier"]],
            "campaign": (values[idx["Campaign"]] or "").strip(),
            "brand": (values[idx["Brand"]] or "").strip(),
            "subcat": (values[idx["Subcategory"]] or "").strip(),
            "imgurl_cur": values[idx["Image URL"]] or "",
            "imgurl_col": idx["Image URL"] + 1,  # 1-based for openpyxl
        })
    return rows


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true", help="report only; no file or xlsx writes")
    ap.add_argument("--category", help="optional: only process this sheet name (e.g. 'Audio & Radio')")
    ap.add_argument("--categories", nargs="+", help="optional: process multiple sheets in one run")
    ap.add_argument("--force", action="store_true", help="re-encode webp even if dst already exists")
    ap.add_argument("--quality", type=int, default=WEBP_QUALITY, help=f"webp quality (default {WEBP_QUALITY})")
    args = ap.parse_args()
    # Override module-level quality from CLI.
    globals()["WEBP_QUALITY"] = args.quality
    # If --force, disable the skip-existing optimization by zeroing the cache check.
    force_reencode = args.force

    if not ORIGINALS.exists():
        print(f"ERROR: {ORIGINALS} not found", file=sys.stderr)
        return 1

    files = sorted(p for p in ORIGINALS.iterdir()
                   if p.is_file() and not p.name.startswith(".")
                   and p.suffix.lower() in {".jpeg", ".jpg", ".png", ".webp", ".avif"})
    print(f"Scanning {len(files)} originals…")

    wb = openpyxl.load_workbook(XLSX)
    rows_by_sheet: dict[str, list[dict]] = {}

    plans = []  # list of (file_path, sheet, matched_rows, web_path)
    unmatched = []
    bad_prefix = []

    for f in files:
        stem = f.stem
        sheet_name, file_key = parse_filename(stem)
        if not sheet_name:
            bad_prefix.append(f.name)
            continue
        if args.category and sheet_name != args.category:
            continue
        if args.categories and sheet_name not in args.categories:
            continue
        if sheet_name not in rows_by_sheet:
            if sheet_name not in wb.sheetnames:
                bad_prefix.append(f"{f.name} -> sheet '{sheet_name}' missing")
                continue
            rows_by_sheet[sheet_name] = load_sheet_rows(wb[sheet_name])
        matched = match_rows(file_key, rows_by_sheet[sheet_name])
        web_path = WEB_DIR / (stem + ".webp")
        rel_url = REL_PREFIX + stem + ".webp"
        if not matched:
            unmatched.append((f.name, sheet_name, file_key))
        plans.append({"file": f, "sheet": sheet_name, "rows": matched,
                      "web_path": web_path, "rel_url": rel_url})

    # -- Report --
    print()
    print(f"=== MATCH REPORT ({'DRY-RUN' if args.dry_run else 'APPLY'}) ===")
    by_sheet = defaultdict(list)
    for p in plans:
        by_sheet[p["sheet"]].append(p)
    for sheet, items in sorted(by_sheet.items()):
        print(f"\n[{sheet}]  ({len(items)} files)")
        for p in items:
            rows = p["rows"]
            if not rows:
                print(f"  ✗ {p['file'].name}  -> no match")
                continue
            tag = f"-> {len(rows)} row(s)" if len(rows) > 1 else "->"
            for r in rows:
                print(f"  ✓ {p['file'].name}  {tag} id={r['id']} {r['tier']} | {r['brand']} / {r['campaign']} ({r['subcat']})")

    if unmatched:
        print(f"\n!! {len(unmatched)} unmatched files:")
        for n, s, k in unmatched:
            print(f"   - {n}  (sheet '{s}', key '{k}')")
    if bad_prefix:
        print(f"\n!! {len(bad_prefix)} files with bad prefix:")
        for n in bad_prefix:
            print(f"   - {n}")

    # -- Plan summary --
    total_rows_to_update = sum(len(p["rows"]) for p in plans)
    total_files_to_convert = len([p for p in plans if p["rows"]])
    print(f"\nSummary: {total_files_to_convert} images to convert, {total_rows_to_update} rows to update")

    if args.dry_run:
        print("\nDRY RUN — no files written, xlsx unchanged.")
        return 0

    if total_files_to_convert == 0:
        print("\nNothing to apply.")
        return 0

    # -- Backup xlsx --
    bak = XLSX.with_name(XLSX.stem + ".bak_before_local_images.xlsx")
    if not bak.exists():
        shutil.copy2(XLSX, bak)
        print(f"\nBackup written: {bak.name}")

    # -- Convert images --
    print("\nConverting images…")
    conv_stats = []
    skipped = 0
    for p in plans:
        if not p["rows"]:
            continue
        orig_kb, new_kb, dim, was_skipped = convert_to_webp(p["file"], p["web_path"], force=force_reencode)
        conv_stats.append((p["file"].name, orig_kb, new_kb, dim))
        if was_skipped:
            skipped += 1

    if conv_stats:
        total_orig = sum(s[1] for s in conv_stats)
        total_new = sum(s[2] for s in conv_stats)
        print(f"   {len(conv_stats)} files ({skipped} skipped/already-converted): "
              f"{total_orig} KB -> {total_new} KB "
              f"({100 * total_new / max(1, total_orig):.0f}% of original)")

    # -- Update xlsx --
    print("\nUpdating xlsx Image URL column…")
    edits = 0
    for p in plans:
        if not p["rows"]:
            continue
        ws = wb[p["sheet"]]
        for r in p["rows"]:
            ws.cell(row=r["row_idx"], column=r["imgurl_col"], value=p["rel_url"])
            edits += 1
    wb.save(XLSX)
    print(f"   {edits} row updates written.")

    print("\nDONE. Next: run cannesdy-site-sync to push the new Image URLs into index.html.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
