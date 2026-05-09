#!/usr/bin/env python3
"""
Surgically update imgUrl fields in index.html based on Image URL column
of the xlsx. Only touches imgUrl lines for rows whose Image URL points
into digital-presentation-images/ (i.e. the local files we just synced).

Match strategy: for each affected xlsx row, find its entry block in
index.html by matching campaign + brand tokens, then rewrite that
block's imgUrl line. Reports any rows it couldn't match.
"""
from __future__ import annotations

import argparse
import json
import re
import shutil
import sys
import unicodedata
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "Brain_Cannesdy_Brand_Experience_2025.xlsx"
HTML = ROOT / "index.html"


def fold_diacritics(s: str) -> str:
    s = unicodedata.normalize("NFKD", s)
    return "".join(ch for ch in s if not unicodedata.combining(ch))


def js_escape(s: str) -> str:
    """Approximate the way the site stores brand/campaign strings as JS literals.

    The DATA block in index.html mostly uses raw printable text and \\u escapes
    for non-ASCII. We try a couple of variants when matching.
    """
    return s.replace("\\", "\\\\").replace('"', '\\"')


def js_unicode_escape(s: str) -> str:
    """Convert non-ASCII chars to JS \\uXXXX form, the way the live HTML stores them."""
    out = []
    for ch in s:
        if ord(ch) < 128:
            out.append(ch)
        else:
            out.append("\\u%04X" % ord(ch))  # uppercase like the file ('\\u00C9')
    return "".join(out)


def make_match_variants(value: str) -> list[str]:
    """Return strings that might appear as the literal in the JS source.

    The HTML stores brand/campaign as JS string literals with non-ASCII as \\uXXXX
    escapes. We generate a few variants — \\uXXXX-escaped (most common in this file),
    raw Unicode, and folded-ASCII — so matching is robust.
    """
    base = js_escape(value)
    out = [js_unicode_escape(base)]
    if base not in out:
        out.append(base)
    folded = js_escape(fold_diacritics(value))
    if folded not in out:
        out.append(folded)
    # Try lowercase variant of the \\uXXXX (some encoders produce \\u00c9 vs \\u00C9)
    upper_form = out[0]
    lower_form = re.sub(r"\\u([0-9A-Fa-f]{4})",
                        lambda m: "\\u" + m.group(1).lower(), upper_form)
    if lower_form not in out:
        out.append(lower_form)
    return out


def find_entry_span(html: str, brand: str, campaign: str) -> tuple[int, int] | None:
    """Find an entry block in DATA whose campaign and brand fields match.

    Returns (start, end) char offsets of the entry's `{...}` block, or None.
    """
    for c_var in make_match_variants(campaign):
        for b_var in make_match_variants(brand):
            # Look for `campaign:"..."` then a brand:"..." in nearby text.
            pattern = re.compile(
                r'\{id:\d+,'                      # entry start
                r'[^{}]*?'                        # tier, etc.
                r'campaign:"' + re.escape(c_var) + r'"'
                r',\s*\n\s*'
                r'brand:"' + re.escape(b_var) + r'"',
                re.MULTILINE,
            )
            m = pattern.search(html)
            if not m:
                continue
            start = m.start()
            # Find the matching closing brace (entries don't nest braces).
            depth = 0
            i = start
            while i < len(html):
                ch = html[i]
                if ch == "{":
                    depth += 1
                elif ch == "}":
                    depth -= 1
                    if depth == 0:
                        return start, i + 1
                i += 1
            return None
    return None


def replace_img_url(block: str, new_url: str) -> tuple[str, bool]:
    """Replace the imgUrl:"..." inside an entry block. Returns (new_block, did_change)."""
    pattern = re.compile(r'(imgUrl:")[^"]*(")')
    if not pattern.search(block):
        return block, False
    new_block, n = pattern.subn(rf'\g<1>{re.escape(new_url)}\g<2>', block, count=1)
    # subn re-escapes specials; undo:
    new_block = new_block.replace(re.escape(new_url), new_url)
    return new_block, n > 0


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(XLSX, data_only=True)
    html = HTML.read_text()

    updates = []  # list of (sheet, brand, campaign, new_url, current_url_in_xlsx)
    for sheet in wb.sheetnames:
        if sheet == "Summary":
            continue
        ws = wb[sheet]
        headers = [c.value for c in ws[1]]
        try:
            i_brand = headers.index("Brand")
            i_camp = headers.index("Campaign")
            i_img = headers.index("Image URL")
        except ValueError:
            continue
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            img = row[i_img] or ""
            if not str(img).startswith("digital-presentation-images/"):
                continue  # only sync rows we just updated
            brand = (row[i_brand] or "").strip()
            campaign = (row[i_camp] or "").strip()
            updates.append((sheet, brand, campaign, str(img)))

    print(f"Found {len(updates)} xlsx rows pointing at local images.")

    new_html = html
    matched = 0
    unmatched = []
    already = 0

    for sheet, brand, campaign, new_url in updates:
        span = find_entry_span(new_html, brand, campaign)
        if span is None:
            unmatched.append((sheet, brand, campaign))
            continue
        block = new_html[span[0]:span[1]]
        if f'imgUrl:"{new_url}"' in block:
            already += 1
            continue
        new_block, ok = replace_img_url(block, new_url)
        if not ok:
            unmatched.append((sheet, brand, campaign))
            continue
        new_html = new_html[:span[0]] + new_block + new_html[span[1]:]
        matched += 1

    print(f"  {matched} entries to update")
    print(f"  {already} already pointing at the new URL")
    print(f"  {len(unmatched)} unmatched")
    if unmatched:
        print("\nUnmatched (entry not found by brand+campaign):")
        for s, b, c in unmatched:
            print(f"   - [{s}] {b} / {c}")

    if args.dry_run:
        return 0

    if matched == 0:
        print("\nNothing to write.")
        return 0

    bak = HTML.with_name(HTML.stem + ".bak_before_local_images.html")
    if not bak.exists():
        shutil.copy2(HTML, bak)
        print(f"\nBackup written: {bak.name}")
    HTML.write_text(new_html)
    print(f"Wrote {matched} imgUrl updates to {HTML.name}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
